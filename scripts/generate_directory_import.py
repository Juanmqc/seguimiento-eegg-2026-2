from __future__ import annotations

import json
import re
import unicodedata
import uuid
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\juan_\OneDrive\Escritorio\Norbert Wiener\ESTUDIOS GENERALES - GESTIÓN\DIRECTORIO EEGG 2026 II LIMA NORTE.xlsx")
PRIVATE = ROOT / "supabase/private-imports/directory-2026-II"
TEACHERS_JSON = PRIVATE / "remote-teachers.json"
COURSES_JSON = PRIVATE / "remote-courses.json"
OUTPUT = PRIVATE / "import_directory.sql"
NS = uuid.UUID("b93cc807-7dd6-42ad-b55e-a7a0e3ff6538")


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm(value: object) -> str:
    value = unicodedata.normalize("NFKD", text(value)).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z0-9]+", " ", value.upper()).strip()


def name_tokens(value: object) -> set[str]:
    return set(norm(value).split())


def source_id(value: object) -> str:
    return text(value).zfill(8)


def sql(value: object) -> str:
    if value is None or value == "":
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def load_rows(path: Path) -> list[dict[str, object]]:
    return json.loads(path.read_text(encoding="utf-8-sig"))["rows"]


def coordinator_course_labels(first: object, second: object) -> list[str]:
    labels: list[str] = []
    for value in (text(first), text(second)):
        if not value:
            continue
        if norm(value) == "CALCULO I II":
            labels.extend(["Cálculo I", "Cálculo II"])
        else:
            labels.append(value)
    return labels


def main() -> None:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    teacher_rows = [r for r in workbook["Plana Docente 2026-I LN"].iter_rows(min_row=2, values_only=True) if r[2]]
    coordinator_rows = [r for r in workbook["Coordinadores 2026-II"].iter_rows(min_row=2, values_only=True) if r[1]]
    remote_teachers = load_rows(TEACHERS_JSON)
    remote_courses = load_rows(COURSES_JSON)

    directory_by_id = {source_id(r[3]): r for r in teacher_rows}
    matches: list[tuple[dict[str, object], tuple[object, ...]]] = []
    for teacher in remote_teachers:
        row = directory_by_id.get(text(teacher["source_identifier"]))
        assert row is not None, f"Docente remoto sin DNI en directorio: {teacher['display_name']}"
        directory_tokens, remote_tokens = name_tokens(row[2]), name_tokens(teacher["display_name"])
        assert directory_tokens <= remote_tokens or remote_tokens <= directory_tokens, f"DNI coincide pero nombre difiere: {row[2]} / {teacher['display_name']}"
        matches.append((teacher, row))
    matched_ids = {text(t["source_identifier"]) for t, _ in matches}
    unmatched_directory = [text(r[2]) for r in teacher_rows if source_id(r[3]) not in matched_ids]
    assert len(remote_teachers) == 17 and len(matches) == 17
    assert len(unmatched_directory) == 1

    juan = next(r for r in teacher_rows if norm(r[2]) == "QUINONEZ COCHACHI JUAN MARCELO")
    assert text(juan[10]).lower() == "juan.quinonez@uwiener.edu.pe"

    course_by_name = {norm(c["name"]): c for c in remote_courses}
    course_aliases = {"ESTILO DE VIDA SALUD Y MEDO AMBIENTE": "ESTILO DE VIDA SALUD Y MEDIO AMBIENTE"}
    coordinators: list[dict[str, object]] = []
    assignments: list[dict[str, object]] = []
    for row in coordinator_rows:
        name, phone, email = text(row[1]), text(row[2]), text(row[4]).lower()
        coordinator_id = str(uuid.uuid5(NS, "coordinator|" + norm(name)))
        coordinators.append({"id": coordinator_id, "name": name, "phone": phone, "email": email})
        for label in coordinator_course_labels(row[5], row[6]):
            normalized = course_aliases.get(norm(label), norm(label))
            course = course_by_name.get(normalized)
            assignments.append({
                "id": str(uuid.uuid5(NS, "assignment|" + norm(name) + "|" + normalized)),
                "coordinator_id": coordinator_id,
                "course_id": course["id"] if course else None,
                "course_name": label,
            })

    matched_course_assignments = sum(a["course_id"] is not None for a in assignments)
    assert len(coordinators) == 8
    assert matched_course_assignments >= 8

    teacher_values = ",\n".join(
        f"  ({sql(t['id'])}::uuid, {sql(text(r[10]).lower())}, {sql(text(r[7]))})" for t, r in matches
    )
    coordinator_values = ",\n".join(
        f"  ({sql(c['id'])}::uuid, {sql(c['name'])}, {sql(c['phone'])}, {sql(c['email'])}, true)" for c in coordinators
    )
    assignment_values = ",\n".join(
        f"  ({sql(a['id'])}::uuid, {sql(a['coordinator_id'])}::uuid, {sql(a['course_id'])}::uuid, {sql(a['course_name'])}, true)" for a in assignments
    )
    content = f"""begin;
select pg_advisory_xact_lock(hashtext('directory-import-2026-II'));

create temp table _teacher_directory (teacher_id uuid, institutional_email text, phone text) on commit drop;
insert into _teacher_directory values
{teacher_values};

update public.teachers t set
  institutional_email = nullif(d.institutional_email, ''),
  phone = nullif(d.phone, ''),
  updated_at = now()
from _teacher_directory d where t.id = d.teacher_id
  and (t.institutional_email, t.phone) is distinct from (nullif(d.institutional_email, ''), nullif(d.phone, ''));

insert into public.course_coordinators (id, full_name, phone, institutional_email, active) values
{coordinator_values}
on conflict (id) do update set full_name=excluded.full_name, phone=excluded.phone,
 institutional_email=excluded.institutional_email, active=true, updated_at=now()
where (course_coordinators.full_name,course_coordinators.phone,course_coordinators.institutional_email,course_coordinators.active)
 is distinct from (excluded.full_name,excluded.phone,excluded.institutional_email,excluded.active);

insert into public.course_coordinator_assignments (id, coordinator_id, course_id, source_course_name, active) values
{assignment_values}
on conflict (id) do update set coordinator_id=excluded.coordinator_id, course_id=excluded.course_id,
 source_course_name=excluded.source_course_name, active=true, updated_at=now()
where (course_coordinator_assignments.coordinator_id,course_coordinator_assignments.course_id,
 course_coordinator_assignments.source_course_name,course_coordinator_assignments.active)
 is distinct from (excluded.coordinator_id,excluded.course_id,excluded.source_course_name,excluded.active);

do $$ begin
  if (select count(*) from _teacher_directory) <> 17 then raise exception 'Se esperaban 17 docentes vinculados'; end if;
  if (select count(*) from public.course_coordinators where active) <> 8 then raise exception 'Se esperaban 8 coordinadores'; end if;
  if (select count(*) from public.course_coordinator_assignments where active and course_id is not null) < 8 then raise exception 'Asignaciones activas insuficientes'; end if;
end $$;
commit;
"""
    PRIVATE.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(content, encoding="utf-8", newline="\n")
    print(f"DOCENTES_DIRECTORIO={len(teacher_rows)}")
    print(f"DOCENTES_VINCULADOS={len(matches)}")
    print("SIN_PROGRAMACION=" + " | ".join(unmatched_directory))
    print(f"COORDINADORES={len(coordinators)}")
    print(f"RELACIONES_COORDINADOR_CURSO={len(assignments)}")
    print(f"RELACIONES_CON_CURSO_ACTIVO={matched_course_assignments}")
    print(f"RELACIONES_SOLO_CONTACTO={len(assignments)-matched_course_assignments}")


if __name__ == "__main__":
    main()
