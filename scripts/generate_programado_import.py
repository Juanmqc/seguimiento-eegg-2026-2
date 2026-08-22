from __future__ import annotations

from collections import Counter
from datetime import datetime, time
from pathlib import Path
import re
import uuid

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT.parent / "ESTUDIOS GENERALES-2026-2.xlsx"
OUTPUT = ROOT / "supabase/private-imports/2026-II/import_programado.sql"
SHEET = "LN_HORARIO_2026-2"
TERM = "2026-II"
NS = uuid.UUID("d960f42d-6d8c-4bbd-b17e-30edb36992a1")


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def sql(value: object) -> str:
    if value is None or value == "":
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def hhmm(value: object) -> str:
    if isinstance(value, (datetime, time)):
        return value.strftime("%H:%M:%S")
    text = clean(value)
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).strftime("%H:%M:%S")
        except ValueError:
            pass
    raise ValueError(f"Hora inválida: {value!r}")


def uid(kind: str, *parts: str) -> str:
    return str(uuid.uuid5(NS, "|".join((kind, *parts))))


def main() -> None:
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb[SHEET]
    rows: list[list[str]] = []
    for values in ws.iter_rows(min_row=4, values_only=True):
        if not any(v is not None and clean(v) for v in values):
            continue
        rows.append([clean(v) for v in values])

    statuses = Counter(r[24].upper() for r in rows)
    assert len(rows) == 175, len(rows)
    assert statuses == {"PROGRAMADO": 136, "CERRADO": 39}, statuses
    active = [r for r in rows if r[24].upper() == "PROGRAMADO"]
    closed_classes = {r[3] for r in rows if r[24].upper() == "CERRADO"}

    records = []
    for r in active:
        original = r[2]
        component = r[4].lower()
        assert component in {"teoría", "práctica"}, component
        main_section = original[:-1] if component == "práctica" and original.endswith("1") else original
        teacher_id = r[20].zfill(8)
        course_code = r[7]
        course_name = r[8]
        teacher_name = r[21]
        class_number = r[3]
        rec = {
            "course_id": uid("course", course_code), "course_code": course_code,
            "course_name": course_name, "area": r[1],
            "teacher_id": uid("teacher", teacher_id), "source_identifier": teacher_id,
            "teacher_name": teacher_name,
            "section_id": uid("section", course_code, main_section, TERM),
            "main_section": main_section, "program": "",
            "assignment_id": uid("assignment", teacher_id, course_code, main_section, TERM),
            "component_id": uid("component", course_code, class_number, original, TERM),
            "original_section": original, "component": component,
            "class_number": class_number, "associated_class": r[5], "class_type": r[6],
            "schedule_id": uid("schedule", course_code, class_number, r[11], hhmm(r[13]), hhmm(r[14]), r[18]),
            "day": r[11].lower(), "start": hhmm(r[13]), "end": hhmm(r[14]),
            "classroom": r[18], "modality": r[10], "shift": r[9],
            "teaching_model": r[12], "academic_hours": r[15], "facility_id": r[17],
            "environment_type": r[19], "environment_capacity": r[16], "teacher_category": r[22],
        }
        assert rec["day"] in {"lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"}
        records.append(rec)

    unique = lambda key: {r[key] for r in records}
    assert len(unique("course_id")) == 11
    assert len(unique("teacher_id")) == 17
    assert len(unique("section_id")) == 85
    assert len(unique("assignment_id")) == 85
    assert len(unique("component_id")) == 136
    assert len(unique("schedule_id")) == 136
    identifiers = unique("source_identifier")
    assert all(re.fullmatch(r"\d{8}", value) for value in identifiers)
    assert sum(value.startswith("0") for value in identifiers) == 4
    assert not (unique("class_number") & closed_classes)
    assert any(r["class_number"] == "1372" and r["day"] == "domingo" for r in records)
    assert {r["class_number"] for r in records if r["course_name"] == "MATEMÁTICA DISCRETA"} == {"1754", "1755", "1895", "1896"}

    by_section: dict[str, list[dict[str, str]]] = {}
    for r in records:
        by_section.setdefault(r["section_id"], []).append(r)
    for r in records:
        if r["component"] == "práctica":
            theories = [t for t in by_section[r["section_id"]] if t["component"] == "teoría"]
            assert theories, f"Práctica sin teoría: {r}"
            assert r["associated_class"] in {t["associated_class"] for t in theories}

    columns = [
        "course_id", "course_code", "course_name", "area", "teacher_id", "source_identifier",
        "teacher_name", "section_id", "main_section", "program", "assignment_id", "component_id",
        "original_section", "component", "class_number", "associated_class", "class_type", "schedule_id", "day",
        "start", "end", "classroom", "modality", "shift", "teaching_model", "academic_hours",
        "facility_id", "environment_type", "environment_capacity", "teacher_category",
    ]
    values = ",\n".join("  (" + ", ".join(sql(r[c]) for c in columns) + ")" for r in records)
    course_codes = ", ".join(sql(x) for x in sorted(unique("course_code")))
    closed_numbers = ", ".join(sql(x) for x in sorted(closed_classes))
    body = f"""\\set ON_ERROR_STOP on
begin;
select pg_advisory_xact_lock(hashtext('academic-import-{TERM}'));

create temp table _academic_programado (
  course_id uuid, course_code text, course_name text, area text,
  teacher_id uuid, source_identifier text, teacher_name text,
  section_id uuid, main_section text, program text, assignment_id uuid,
  component_id uuid, original_section text, component text, class_number text,
  associated_class text, class_type text, schedule_id uuid, day text, start time, \"end\" time,
  classroom text, modality text, shift text, teaching_model text, academic_hours numeric,
  facility_id text, environment_type text, environment_capacity integer, teacher_category text
) on commit drop;

insert into _academic_programado values
{values};

do $$ begin
  if (select count(*) from _academic_programado) <> 136 then raise exception 'Se esperaban 136 filas'; end if;
  if (select count(distinct course_id) from _academic_programado) <> 11 then raise exception 'Cursos inválidos'; end if;
  if (select count(distinct teacher_id) from _academic_programado) <> 17 then raise exception 'Docentes inválidos'; end if;
  if (select count(distinct section_id) from _academic_programado) <> 85 then raise exception 'Secciones inválidas'; end if;
end $$;

insert into public.courses (id, code, name, area, active)
select distinct course_id, course_code, course_name, area, true from _academic_programado
on conflict (id) do update set code=excluded.code, name=excluded.name, area=excluded.area, active=true,
 updated_at=now() where (courses.code,courses.name,courses.area,courses.active) is distinct from
 (excluded.code,excluded.name,excluded.area,excluded.active);

insert into public.teachers (id, profile_id, display_name, institutional_email, source_identifier, active)
select distinct teacher_id, null::uuid, teacher_name, null::text, source_identifier, true from _academic_programado
on conflict (id) do update set display_name=excluded.display_name, source_identifier=excluded.source_identifier, active=true,
 updated_at=now() where (teachers.display_name,teachers.source_identifier,teachers.active) is distinct from
 (excluded.display_name,excluded.source_identifier,excluded.active);

insert into public.sections (id, course_id, section_code, academic_program, modality, campus, academic_term, active)
select distinct section_id, course_id, main_section, null::text, modality, 'Lima Norte', '{TERM}', true
from _academic_programado
on conflict (id) do update set course_id=excluded.course_id, section_code=excluded.section_code,
 academic_program=excluded.academic_program, modality=excluded.modality, campus=excluded.campus,
 academic_term=excluded.academic_term, active=true, updated_at=now()
where (sections.course_id,sections.section_code,sections.academic_program,sections.modality,sections.campus,sections.academic_term,sections.active)
 is distinct from (excluded.course_id,excluded.section_code,excluded.academic_program,excluded.modality,excluded.campus,excluded.academic_term,excluded.active);

insert into public.teacher_assignments (id, teacher_id, section_id, academic_term, active, teacher_category)
select distinct assignment_id, teacher_id, section_id, '{TERM}', true, teacher_category
from _academic_programado
on conflict (id) do update set teacher_id=excluded.teacher_id, section_id=excluded.section_id,
 academic_term=excluded.academic_term, active=true, teacher_category=excluded.teacher_category
where (teacher_assignments.teacher_id,teacher_assignments.section_id,teacher_assignments.academic_term,
 teacher_assignments.active,teacher_assignments.teacher_category)
 is distinct from (excluded.teacher_id,excluded.section_id,excluded.academic_term,excluded.active,excluded.teacher_category);

insert into public.section_components (id, section_id, original_section_code, component,
 class_number, associated_class, class_type)
select distinct component_id, section_id, original_section, component::public.class_component,
 class_number::bigint, nullif(associated_class,'')::integer, class_type from _academic_programado
on conflict (id) do update set section_id=excluded.section_id,
 original_section_code=excluded.original_section_code, component=excluded.component, class_number=excluded.class_number,
 associated_class=excluded.associated_class, class_type=excluded.class_type, updated_at=now()
where (section_components.section_id,section_components.original_section_code,
 section_components.component,section_components.class_number,section_components.associated_class,section_components.class_type)
 is distinct from (excluded.section_id,excluded.original_section_code,
 excluded.component,excluded.class_number,excluded.associated_class,excluded.class_type);

insert into public.schedules (id, teacher_assignment_id, section_component_id, day_of_week, start_time, end_time,
 classroom, modality, shift, teaching_model, academic_hours, facility_id, environment_type, environment_capacity)
select distinct schedule_id, assignment_id, component_id, day::public.weekday, start, \"end\", classroom, modality,
 shift, teaching_model, academic_hours, facility_id, environment_type, environment_capacity
from _academic_programado
on conflict (id) do update set teacher_assignment_id=excluded.teacher_assignment_id,
 section_component_id=excluded.section_component_id, day_of_week=excluded.day_of_week,
 start_time=excluded.start_time, end_time=excluded.end_time, classroom=excluded.classroom,
 modality=excluded.modality, shift=excluded.shift, teaching_model=excluded.teaching_model,
 academic_hours=excluded.academic_hours, facility_id=excluded.facility_id,
 environment_type=excluded.environment_type, environment_capacity=excluded.environment_capacity, updated_at=now()
where (schedules.teacher_assignment_id,schedules.section_component_id,schedules.day_of_week,schedules.start_time,
 schedules.end_time,schedules.classroom,schedules.modality,schedules.shift,schedules.teaching_model,
 schedules.academic_hours,schedules.facility_id,schedules.environment_type,schedules.environment_capacity) is distinct from
 (excluded.teacher_assignment_id,excluded.section_component_id,excluded.day_of_week,excluded.start_time,
 excluded.end_time,excluded.classroom,excluded.modality,excluded.shift,excluded.teaching_model,
 excluded.academic_hours,excluded.facility_id,excluded.environment_type,excluded.environment_capacity);

do $$ begin
  if (select count(*) from public.courses where code in ({course_codes})) <> 11 then raise exception 'No se cargaron 11 cursos'; end if;
  if (select count(*) from public.teachers where id in (select teacher_id from _academic_programado)) <> 17 then raise exception 'No se cargaron 17 docentes'; end if;
  if (select count(*) from public.sections where id in (select section_id from _academic_programado)) <> 85 then raise exception 'No se cargaron 85 secciones'; end if;
  if (select count(*) from public.teacher_assignments where id in (select assignment_id from _academic_programado)) <> 85 then raise exception 'No se cargaron 85 asignaciones'; end if;
  if (select count(*) from public.section_components where id in (select component_id from _academic_programado)) <> 136 then raise exception 'No se cargaron 136 componentes'; end if;
  if (select count(*) from public.schedules where id in (select schedule_id from _academic_programado)) <> 136 then raise exception 'No se cargaron 136 horarios'; end if;
  if exists (select 1 from public.section_components where class_number in ({closed_numbers})) then raise exception 'Se cargó una clase CERRADO'; end if;
end $$;

commit;
\\echo ACADEMIC_IMPORT_2026_II_OK
"""
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(body, encoding="utf-8", newline="\n")
    print(f"Generado: {OUTPUT}")
    print("Validación fuente: 175 total, 136 PROGRAMADO, 39 CERRADO; 11/17/85/85/136/136")


if __name__ == "__main__":
    main()
